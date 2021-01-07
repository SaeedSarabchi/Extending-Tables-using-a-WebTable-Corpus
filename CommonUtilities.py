'''
    This file contains the common utility functions used throughout the code-base.
'''

import json
import pickle
import sys
import os
import copy
import editdistance 
import csv
from unidecode import  unidecode
from urllib.parse import unquote
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
import openpyxl
from openpyxl.styles import Font, Fill
from openpyxl.styles.borders import Border, Side
import ast #ast.literal_eval 


# Dataset Choices:
# Basketball
# Camera_Brand
# US_State_Governors
# Hockey
#Movie_OR_Film
#Film_Director
#Music_Album
#T2D_233
#TestDataSets/Test1
#TestDataSets/TPC-H
#TestDataSets/T2D_233
#TestDataSets/currency_dataset
#TestDataSets/city_dataset

Dataset = "TestDataSets/T2D_233"
#Dataset = "T2D_233"
####### GENERAL CONSTANTS #########
def get_Constant(field):
    

    
    Constants={}

    Constants["Json_Files_Directory"] =   'C:/Saeed_Local_Files/'+str(Dataset)
    Constants["RelatedTables"] =   'C:/Saeed_Local_Files/RelatedTables/'+str(Dataset)
    
    Constants["Merged_Json_URL"] =        Constants["Json_Files_Directory"]+"/Merged.json"
    Constants["Filter_Json_URL"] =        Constants["Json_Files_Directory"]+"/Filtered.json"
    Constants["Deduplicated_Merged_Json_URL"] =        Constants["Json_Files_Directory"]+"/Deduplicated_Merged.json"
    Constants["Deduplicated_Merged_Pickle_URL"] =        Constants["Json_Files_Directory"]+"/Deduplicated_Merged.pickle"
    Constants["Stitched_Json_URL"] =        Constants["Json_Files_Directory"]+"/Stitched.json"
    Constants["EAB_Merged_Json_URL"]=     Constants["Json_Files_Directory"]+'/EAB_Merged.json'
    Constants["EAB_Merged_Pickle_URL"] =  Constants["Json_Files_Directory"]+'/EAB_Merged.pickle'
    Constants["WIK_Pickle_URL"] =         Constants["Json_Files_Directory"]+'/WIK_Index.pickle'
    Constants["WIA_Pickle_URL"] =         Constants["Json_Files_Directory"]+'/WIA_Index.pickle'
    Constants["Feature_URL_Pickle"] = Constants["Json_Files_Directory"] + '/Feature_URL_Pickle.pickle'
    Constants["EAB_Small_Pickle"] = Constants["Json_Files_Directory"] + '/EAB_Small_Pickle.pickle'
    Constants["EAB_Label_Pickle"] = Constants["Json_Files_Directory"] + '/EAB_Label_Pickle.pickle'
    Constants["SMW_Graph_Pickle"] = Constants["Json_Files_Directory"] + '/SMW_Graph_Pickle.pickle'
    Constants["Feature_CONTEXT_Pickle"] = Constants["Json_Files_Directory"] + '/Feature_CONTEXT_Pickle.pickle'
    Constants["Feature_TUPLE_Pickle"] = Constants["Json_Files_Directory"] + '/Feature_TUPLE_Pickle.pickle'
    Constants["Feature_TableContent_Pickle"] = Constants["Json_Files_Directory"] + '/Feature_TableContent_Pickle.pickle'
    Constants["Feature_Table2Context_Pickle"] = Constants["Json_Files_Directory"] + '/Feature_Table2Context_Pickle.pickle'
    Constants["Feature_AttributeNames_Pickle"] = Constants["Json_Files_Directory"] + '/Feature_AttributeNames_Pickle.pickle'
    Constants["Feature_ColValues_Pickle"] = Constants["Json_Files_Directory"] + '/Feature_ColValues_Pickle.pickle'
    Constants["Feature_ColWidth_Pickle"] = Constants["Json_Files_Directory"] + '/Feature_ColWidth_Pickle.pickle'
    Constants["T2Syn_DMA_Pickle"] = Constants["Json_Files_Directory"] + '/T2Syn_DMA_Pickle.pickle'
    Constants["Old_EAB_Merged_Json_URL"] = Constants["Json_Files_Directory"] + '/Old_EAB_Merged_Json_URL.json'
    Constants["FPPR"] = Constants["Json_Files_Directory"] + '/FPPR.pickle'
    Constants["T2Syn_Direct"] = Constants["Json_Files_Directory"] + '/T2Syn_Direct.pickle'
    Constants["T2Syn_From_SMW"] = Constants["Json_Files_Directory"] + '/T2Syn_From_SMW.pickle'
    Constants["GroundTruth"] = Constants["Json_Files_Directory"] + '/GroundTruth.xlsx'
    Constants["Experiments"] = Constants["Json_Files_Directory"] + '/Experiments.xlsx'
    Constants["Search_GroundTruth_In_Dataset"] = Constants["Json_Files_Directory"] + '/Search_GroundTruth_In_Dataset.xlsx'
    Constants["RelatedTable_WIK_Pickle_URL"] =         Constants["RelatedTables"]+'/WIK_Index.pickle'
    Constants["RelatedTable_WIK_NoFreeBase_Pickle_URL"] =         Constants["RelatedTables"]+'/WIK_NoFreeBase_Index.pickle'
    
    Constants["Header_ACSDB_Pickle_URL"] =    Constants["Json_Files_Directory"]+'/Header_ACSDB_Pickle.pickle'
    Constants["Content_ACSDB_Pickle_URL"] =    Constants["Json_Files_Directory"]+'/Content_ACSDB_Pickle.pickle'
    Constants["ColumnPostingList_Pickle_URL"] =    Constants["Json_Files_Directory"]+'/ColumnPostingList_Pickle_URL.pickle'
    Constants["JsonTablesFolder"] =    Constants["Json_Files_Directory"]+'/json_tables'
    Constants["CsvTablesFolder"] =    Constants["Json_Files_Directory"]+'/csv_tables'
    Constants["OutputFolder"] =    Constants["Json_Files_Directory"]+'/output'
    Constants["OutputFolder2"] =    Constants["Json_Files_Directory"]+'/output2'
    
    



    if field not in Constants:
        raise ValueError ("Error: Field Not in Constants!!")


    return Constants[field]


####### UTILITY FUNCTIONS #########
def string_Cleanse(string):
    return str(string).strip().lower().replace(" ", "").replace("-", "").\
        replace("/", "").replace("(", "").replace(")", "")\
        .replace("_", "").replace("]","").replace("[","").replace(".","").replace(":","").replace("'","")
        
        
def Merge_Tables(table1, table2):
    MergedTables=[]
    for i in range(len(table1[0])):
        row=[]
        for j in range(len(table1)):
            row.append(table1[j][i])
        MergedTables.append(row)
        
    for i in range(len(table2[0])):
        row=[]
        for j in range(len(table2)):
            row.append(table2[j][i])
        MergedTables.append(row)
        
    return MergedTables

def TableContent_To_OrderedStr(table):
    Result_array=[]
    for i in range(len(table[0])):
        for j in range(len(table)):
            Result_array.append(string_Cleanse(table[j][i]))
    
    Result_Str=""
    Sorted_Array =  sorted(Result_array, key=lambda x:str(x))
    for elem in Sorted_Array:
        Result_Str+=str(elem)
    return Result_Str


def TableContent_To_Str(table):
    Result_String=""
    for i in range(len(table[0])):
        for j in range(len(table)):
            Result_String+=((table[j][i]))

    return Result_String
     
        
def JsonTable_To_String(jsonTable):
    table=[]
    keyCol=-1
    URL=""
    Context=""
    tableNum=""
    s3Link=""
    recordOffset=""
    pageTitle=""
    title=""
    
    try:
        table = jsonTable["relation"]
        keyCol = jsonTable["keyColumnIndex"]    
        URL = jsonTable["url"]
        Context = jsonTable["textBeforeTable"]+" "+jsonTable["textAfterTable"]
        tableNum = jsonTable["tableNum"]
        s3Link= jsonTable["s3Link"]
        recordOffset= jsonTable["recordOffset"]
        pageTitle = jsonTable["pageTitle"]
        title = jsonTable["title"]
    
    except:
        pass
        
    HeaderIndex = -1
    if(jsonTable["hasHeader"]):
        HeaderIndex = jsonTable["headerRowIndex"]
    
    Result="\n"
    for i in range(len(table[0])):
        if(i == HeaderIndex):
                Result += "@ "
        for j in range(len(table)):
            
            if(j==keyCol):
                Result += " *["+str(table[j][i])+"]* "
            else:
                Result += " ["+str(table[j][i])+"] "
        if(i == HeaderIndex):
           Result += " @"
        Result += "\n"
    Result +="****\n"

    Result += "\npageTitle: "+ str(pageTitle)
    Result += "\ntitle: "+ str(title)
    Result += "\nContext: "+ str(Context) + "\n"
    Result += "\nURL: "+ str(URL) + "\n"
    Result += "tableNum: "+ str(tableNum) + "\n"
    Result += "s3Link: "+ str(s3Link) + "\n"
    Result += "recordOffset: "+ str(recordOffset) + "\n"
    Result +="----------------------\n"
    return Result


def TableContent_To_TabledStr(table):
    Result=""
    for i in range(len(table[0])):
        for j in range(len(table)):
            

            Result += '{: <30}'.format(" ["+str(table[j][i])+"] ")

        Result += "\n"
    Result +="****\n\n"

    return Result



def Edit_Distance_Ratio(str1,str2):
    
    #return ((1- fuzz.token_set_ratio (str1, str2)/100))
    return ((1- fuzz.token_set_ratio (str1, str2)/100))


def Char_Level_Edit_Distance(str1,str2):
    
    #return ((1- fuzz.token_set_ratio (str1, str2)/100))
    return ((1- fuzz.ratio (string_Cleanse(str1), string_Cleanse(str2))/100))


def Write_Dict_W_Table_To_File_Generic(Dict,DictName, SaveFolderLocation, TableType):
    dict={}
    if (TableType=="EAB"):
        with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
            dict = pickle.load(f3)
            
    if (TableType=="Orig_Table"):
        with open(get_Constant("Deduplicated_Merged_Pickle_URL"), 'rb') as f3:
            dict = pickle.load(f3)
            
        
    
    Textfile = open(
            SaveFolderLocation + "/"+Dataset +"_"+DictName+".txt", "w", encoding="utf-8")
    sorted_x = sorted(Dict.items(), key=lambda item:item[1], reverse=True)
    line_Num=0
    for data in sorted_x:
        line_Num+=1
        Textfile.write(repr(line_Num)+"-Node: "+repr(data[0])+" : "+repr(data[1])+"\n")
        Textfile.write(JsonTable_To_String(dict[data[0]]))
        Textfile.write("##############################################\n")
        
        
    
    Textfile.close()
    
    
    
    
def Write_Only_Dict_To_File_Generic(Dict, Name, SaveFolderLocation):
    Textfile = open(
                SaveFolderLocation +"/"+Dataset +"_"+Name+".txt", "w", encoding="utf-8")
    line_Num=0
    for key in Dict:
     
        Textfile.write(repr(key)+" :\n"+repr(Dict[key])+"\n")

        Textfile.write("##############################################\n")
        
        
    
    Textfile.close()
    
    

def Write_Table_To_CSV(Table, outFile):
    #print("###Write_Table_To_CSV(Start):####")
    #print(Table)
    with open(outFile, "w",encoding='utf8',newline='') as csvFile:
        writer = csv.writer(csvFile, quoting=csv.QUOTE_ALL)
        
        
        for i in range(len(Table[0])):
            writeRow=[]
            for j in range(len(Table)):
                decodedString=""
                if(Table[j][i]!=None):
                    decodedString=""
                    try:
                        decodedString = unidecode(unquote(Table[j][i]))   
                    except:
                        decodedString = str(Table[j][i])
                writeRow.append(decodedString)
            writer.writerow(writeRow)  
    
    #print("###Write_Table_To_CSV(End):####")        
    #writer.close()


#If Value is contained in the set, based on edit distance with the input threshold,
#Return the value (of the set), else, return ""
def Approximate_Match_In_Set(Value,Set, Threshold):
    
    for SetVal in Set:
        if(Edit_Distance_Ratio(Value, SetVal)<Threshold):
            return SetVal
    
    return "";
    
    
def Populate_QueryKeys_From_ExcelFile(ExcelFile):
    wb = openpyxl.load_workbook(ExcelFile)
    sheet = wb.worksheets[0]
    Input_Keys=[]
    for i in range(2, sheet.max_row+1):
        Input_Keys.append(sheet.cell(row=i, column=1).value)
    
    return Input_Keys 

def Populate_QueryTable_From_ExcelFile(ExcelFile):
    wb = openpyxl.load_workbook(ExcelFile)
    sheet = wb.worksheets[0]
    Input_Table=[]
    for j in range(1, sheet.max_column+1):
        Input_Keys=[]
        for i in range(2, sheet.max_row+1):
            Input_Keys.append(sheet.cell(row=i, column=j).value)
        Input_Table.append(Input_Keys)
    
    return Input_Table 

def csvTables_to_jsonFiles(csvTableFolder, jsonFileFolder):    
    '''
        converts each table in csv to a json file
    '''
    
    for csvFile in os.listdir(csvTableFolder):
        with open(csvTableFolder+"/"+csvFile, encoding='utf8') as csvTable:
            csv_Content = csv.reader(csvTable)
            table = []
            for row in csv_Content:
                
                #print("csvFile:"+str(csvFile))
                #print("row:"+str(row))
                
                #initiliaze table columns with the headers:
                if(len(table)==0):
                    for val in row:
                        table.append([val])
                
                else:
                    #populate the table's columns
                    for i in range(len(table)):
                        if(len(row)>i):
                            table[i].append(row[i])
                        else:
                            table[i].append("")
                            
                    
            #Populate json file:
            jsonDoc={"headerRowIndex": 0,
                      "hasHeader": True, 
                      "relation": table
                    }
            
            #Save the JsonFile
            with open(jsonFileFolder+"/"+csvFile[:-4]+".json", "w") as jsonFile:
                json.dump(jsonDoc, jsonFile)
            
                        
                
                        
def Convert_CSVFile_To_Table(csvFilePath):
    with open(csvFilePath, encoding='utf8') as csvFile:
        csvReader = csv.reader(csvFile, delimiter=',')
        table=[]
        Headers = next(csvReader)
        for header in Headers:
            #table.append([header])
            if(len(header)>0):
                if(header[0]=="["):
                        table.append([ast.literal_eval(header)])
                else:
                        table.append([header])
            else:
                table.append([header])
        
        for line in csvReader:
            for i in range(len(line)):
                if(len(line[i])>0):
                    if(line[i][0]=="["):
                        table[i].append(ast.literal_eval(line[i]))
                    else:
                        table[i].append(line[i])
                else:
                    table[i].append(line[i])
                
        return table
    
def Convert_CSV_TO_XLSX(csvFilePath):
    with open(csvFilePath, encoding='utf8') as csvFile:
        csvReader = csv.reader(csvFile, delimiter=',')
       
        workbook  = openpyxl.Workbook()
        worksheet = workbook.active
        
        thin_border = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))
        
        
        #mono_spaced_font = openpyxl.styles.Font(name='Courier New', size=10, bold=False)
        
        i=-1
        for line in (csvReader):
            i+=1
            for j in range(len(line)):
                print(line[j])
                worksheet.cell(row=i+1, column=j+1).value = line[j]
                worksheet.cell(row=i+1, column=j+1).alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center', wrapText='True')
                #worksheet.cell(row=i+1, column=j+1).border = thin_border
                #worksheet.cell(row=i+1, column=j+1).font = mono_spaced_font
                
                
                
        #From https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column # Get the column name
            for cell in col:
                 try: # Necessary to avoid error on empty cells
                     if len(str(cell.value)) > max_length:
                         max_length = len(cell.value)
                 except:
                     pass
            adjusted_width = (max_length+2) * 1.2 
            worksheet.column_dimensions[column].width = adjusted_width
                
        
    
        workbook.save(csvFilePath[:-4] + '.xlsx')
        

def clean_csv_file(csvTableFolder):    
    '''
        clean each table in csv : replace all &nbsp
    '''
    
    for csvFile in os.listdir(csvTableFolder):
        table = Convert_CSVFile_To_Table(csvTableFolder+"/"+csvFile)
        for i in range(len(table)):
            for j in range(len(table[0])):
                table[i][j]=table[i][j].replace("&nbsp;"," ")
        
            
        #Save the csvFile
        #print(table)
        Write_Table_To_CSV(table,csvTableFolder+"/"+csvFile)
            
                

    