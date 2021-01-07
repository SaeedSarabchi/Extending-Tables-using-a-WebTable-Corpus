import json
import pickle
import copy
from WDC_Analysis import *
#from scipy import spatial
#from sklearn.metrics.pairwise import cosine_similarity
#from sklearn.feature_extraction.text import TfidfVectorizer
import collections
from CommonUtilities import *
from SMW_Graph_Module import *
from Dataset_Gathering import *
#from sklearn import svm
#from sklearn import tree
#from sklearn.naive_bayes import GaussianNB
import math
#import numpy as np
import re, math
from collections import Counter
import os
import openpyxl
from openpyxl.styles import Font, Fill
from CommonUtilities import *
import csv
import Octopus

json_obj1 = []
json_obj2 = []



Octopus_Value_Dist_Threshold  = 0.3

def Populate_QueryKeys():
    wb = openpyxl.load_workbook(get_Constant("GroundTruth"))
    sheet = wb.get_sheet_by_name('GroundTruth')
    Input_Keys=[]
    for i in range(2, sheet.max_row+1):
        Input_Keys.append(sheet.cell(row=i, column=1).value)
    
    return Input_Keys
        
    
    
def DMA_Attribute_Discovery(keyList, Threshold, Synonym_Enabled=False):
    keyList2=[]
    for x in range(len(keyList)):
        keyList2.append(string_Cleanse(keyList[x]))

    queryKeys=[]
    for item in keyList2:
        if (item not in queryKeys ):
            queryKeys.append(item)

    with open(get_Constant("WIK_Pickle_URL"), 'rb') as f3:
        WIK = pickle.load(f3)
        DMA_MATCH={}
        for queryKey in queryKeys:
            if(queryKey in WIK):
                WIK_Tables = WIK[queryKey]
                for table in WIK_Tables:
                    if (table not in DMA_MATCH):
                        DMA_MATCH[table] = 1
                    else:
                        DMA_MATCH[table] += 1

        Synonym_Indx={}
        synonyms=[]
        tfidf_vectorizer = TfidfVectorizer(use_idf=False, smooth_idf=False)


        if(Synonym_Enabled):
            with open(get_Constant("T2Syn_DMA_Pickle"), 'rb') as f3:
                Synonym_Indx = pickle.load(f3)
        with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
            WT_Corpus = pickle.load(f3)
            DMA_Result={}
            querySize=len(queryKeys)
            Attribute_Name_Cleanse_Map={}
            for table in DMA_MATCH:
                minsize =min(querySize, (len(WT_Corpus[table]["relation"][0])-1))
                DMA_Score=(DMA_MATCH[table]/minsize)
                if(DMA_Score>=Threshold):
                    headerIndex = WT_Corpus[table]["headerRowIndex"]
                    Attribute_Name = WT_Corpus[table]["relation"][1][headerIndex]
                    Key_Name = WT_Corpus[table]["relation"][0][headerIndex]
                    DMA_Result[table] = (DMA_Score,string_Cleanse(Attribute_Name), Key_Name )
                    Attribute_Name_Cleanse_Map[string_Cleanse(Attribute_Name)]=Attribute_Name
                    if(Synonym_Enabled):
                        if (table in Synonym_Indx):
                            syns = Synonym_Indx[table]
                            for syn_item in syns:
                                if (syn_item[1] > Threshold):
                                    documents = []
                                    documents.append(
                                        WT_Corpus[table]["textBeforeTable"] + WT_Corpus[table]["textAfterTable"])
                                    documents.append(WT_Corpus[syn_item[2]]["textBeforeTable"] + WT_Corpus[syn_item[2]][
                                        "textAfterTable"])
                                    tfidf = tfidf_vectorizer.fit_transform(documents)
                                    # print (tfidf_matrix.shape)
                                    pairwise_sim_sparse = tfidf * tfidf.T
                                    pairwise_similarity = pairwise_sim_sparse.toarray()
                                    if (pairwise_similarity[0][1] < 0.95):
                                        # if(WT_Corpus[tableKey]["url"]!= WT_Corpus[syn_item[2]]["url"]):
                                        synonyms.append((table, syn_item[0], syn_item[2], syn_item[1]))
                '''
                for table in DMA_Result:
                    print(table+":"+repr(DMA_Result[table]))
                '''
            print("\n ################ ATTRIBUTE DISCOVERY ############### \n")

            AD=[]
            for table in DMA_Result:
                    if DMA_Result[table][1] not in AD:
                        AD.append(DMA_Result[table][1])



            Key_Names=[]
            for table in DMA_Result:
                    if DMA_Result[table][2] not in Key_Names:
                        Key_Names.append(DMA_Result[table][2])

            return_AD=[]
            for attrib in AD:
                return_AD.append(Attribute_Name_Cleanse_Map[attrib])

            for syn in synonyms:
                if syn[1] not in return_AD:
                    return_AD.append(syn[1])

            print("Number of Augmented Attributes: " + repr(len(return_AD)))
            print ("Attributes Discoverd for Augmentation: "+ repr(return_AD))
            #print ("###########################")
            #print ("Attribute Names used for the query KEYs in webtables, Name(Q.K): " + str(Key_Names))
            #print ("###########################")
            return return_AD


def DMA_MATCH(keyList, Aug_Attribute_Name, Threshold ):
    keyListMap = {}
    keyList2 = []
    # Cleansing the Key List + Unique
    for x in range(len(keyList)):
        keyListMap[string_Cleanse(keyList[x])] = keyList[x]
        keyList2.append(string_Cleanse(keyList[x]))

    queryKeys = []
    for item in keyList2:
        if (item not in queryKeys):
            queryKeys.append(item)

    # Cleansing the Attribute Name
    Orig_Aug_Attribute_Name = Aug_Attribute_Name
    Aug_Attribute_Name = string_Cleanse(Aug_Attribute_Name)


    DMA_MATCH = {}
    with open(get_Constant("WIK_Pickle_URL"), 'rb') as f3:
        WIK = pickle.load(f3)
        DMA_MATCH = {}
        for queryKey in queryKeys:
            if (queryKey in WIK):
                WIK_Tables = WIK[queryKey]
                for table in WIK_Tables:
                    if (table not in DMA_MATCH):
                        DMA_MATCH[table] = 1
                    else:
                        DMA_MATCH[table] += 1

        # Populating DMA_Result(tableKey->(DMA Score,Related Attribute))
        with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
            WT_Corpus = pickle.load(f3)
            DMA_Result = {}
            querySize = len(queryKeys)
            for table in DMA_MATCH:
                minsize = min(querySize, (len(WT_Corpus[table]["relation"][0]) - 1))
                DMA_Score = (DMA_MATCH[table] / minsize)
                if (DMA_Score >= Threshold):
                    tableID = WT_Corpus[table]["ID"]
                    headerIndex = WT_Corpus[table]["headerRowIndex"]
                    Attribute_Name = WT_Corpus[table]["relation"][1][headerIndex]
                    DMA_Result[table] = (DMA_Score, Attribute_Name, tableID)

            return DMA_Result
        
        
        
        
def AD_Seed_Tables(keyList, Threshold ):
    keyListMap = {}
    keyList2 = []
    # Cleansing the Key List + Unique
    for x in range(len(keyList)):
        keyListMap[string_Cleanse(keyList[x])] = keyList[x]
        keyList2.append(string_Cleanse(keyList[x]))

    queryKeys = []
    for item in keyList2:
        if (item not in queryKeys):
            queryKeys.append(item)


    DMA_MATCH = {}
    with open(get_Constant("WIK_Pickle_URL"), 'rb') as f3:
        WIK = pickle.load(f3)
        DMA_MATCH = {}
        for queryKey in queryKeys:
            if (queryKey in WIK):
                WIK_Tables = WIK[queryKey]
                for table in WIK_Tables:
                    if (table not in DMA_MATCH):
                        DMA_MATCH[table] = 1
                    else:
                        DMA_MATCH[table] += 1

        # Populating DMA_Result(tableKey->(DMA Score,Related Attribute))
        with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
            WT_Corpus = pickle.load(f3)
            DMA_Result = {}
            querySize = len(queryKeys)
            for table in DMA_MATCH:
                minsize = min(querySize, (len(WT_Corpus[table]["relation"][0]) - 1))
                DMA_Score = (DMA_MATCH[table] / minsize)
                if (DMA_Score >= Threshold):
                    tableID = WT_Corpus[table]["ID"]
                    headerIndex = WT_Corpus[table]["headerRowIndex"]
                    Attribute_Name = WT_Corpus[table]["relation"][1][headerIndex]
                    #DMA_Result[table] = (DMA_Score, Attribute_Name, tableID)
                    DMA_Result[table] = (DMA_Score)

            #Normalize DMA Scores:
            '''Total_Score=0
            for table in DMA_Result:
                Total_Score+=DMA_Result[table]
            
            for table in DMA_Result:
                DMA_Result[table] = DMA_Result[table]/Total_Score'''
                
            
            return DMA_Result




def AD_Table_Search(keyList, Threshold, MatchType):
    Cleansed_queryKeys_set = set()
    # Cleansing the Key List + Unique
    for key in keyList:
        Cleansed_queryKeys_set.add(string_Cleanse(key))


    if(MatchType=="Approximate_Match"):
        EditDistanceThreshold = 0.3
        

        
    DMA_MATCH = {}
    with open(get_Constant("WIK_Pickle_URL"), 'rb') as f3:
        WIK = pickle.load(f3)
        
        if(MatchType=="Exact_Match"):
            DMA_MATCH = {}
            for queryKey in Cleansed_queryKeys_set:
                if (queryKey in WIK):
                    WIK_Tables = WIK[queryKey]
                    for table in WIK_Tables:
                        if (table not in DMA_MATCH):
                            DMA_MATCH[table] = 1
                        else:
                            DMA_MATCH[table] += 1
                            
        if(MatchType=="Approximate_Match"):
            DMA_MATCH = {}
            for queryKey in Cleansed_queryKeys_set:
                TableSet= set()
                for key in WIK:
                    if(Edit_Distance_Ratio(key, queryKey)<EditDistanceThreshold):
                        WIK_Tables = WIK[key]
                        for table in WIK_Tables:
                            TableSet.add(table)
                        
                        
                        
                for table in TableSet:
                    if (table not in DMA_MATCH):
                        DMA_MATCH[table] = 1
                    else:
                        DMA_MATCH[table] += 1
                
                
                
            

        # Populating DMA_Result(tableKey->(DMA Score,Related Attribute))
        with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
            WT_Corpus = pickle.load(f3)
            DMA_Result = {}
            querySize = len(Cleansed_queryKeys_set)
            for table in DMA_MATCH:
                KeyColSize = len(set(WT_Corpus[table]["relation"][0]))
                if(WT_Corpus[table]["hasHeader"]):
                    KeyColSize-=1
                minsize = min(querySize, KeyColSize)
                DMA_Score = (DMA_MATCH[table] / minsize)
                if (DMA_Score >= Threshold):
                    DMA_Result[table] = (DMA_Score)



            return DMA_Result



def ABA_Seed_Tables(keyList, AtrName, Threshold):
    keyListMap = {}
    keyList2 = []
    # Cleansing the AtrName Key List + Unique
    AtrName =  string_Cleanse(AtrName)
    
    for x in range(len(keyList)):
        keyListMap[string_Cleanse(keyList[x])] = keyList[x]
        keyList2.append(string_Cleanse(keyList[x]))

    queryKeys = []
    for item in keyList2:
        if (item not in queryKeys):
            queryKeys.append(item)


    DMA_MATCH = {}
    with open(get_Constant("WIK_Pickle_URL"), 'rb') as f3:
        WIK = pickle.load(f3)
        DMA_MATCH = {}
        for queryKey in queryKeys:
            if (queryKey in WIK):
                WIK_Tables = WIK[queryKey]
                for table in WIK_Tables:
                    if (table not in DMA_MATCH):
                        DMA_MATCH[table] = 1
                    else:
                        DMA_MATCH[table] += 1

        # Populating DMA_Result(tableKey->(DMA Score,Related Attribute))
        with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
            WT_Corpus = pickle.load(f3)
            DMA_Result = {}
            querySize = len(queryKeys)
            for table in DMA_MATCH:
                minsize = min(querySize, (len(WT_Corpus[table]["relation"][0]) - 1))
                DMA_Score = (DMA_MATCH[table] / minsize)
                if (DMA_Score >= Threshold):
                    tableID = WT_Corpus[table]["ID"]
                    headerIndex = WT_Corpus[table]["headerRowIndex"]
                    Attribute_Name = string_Cleanse(WT_Corpus[table]["relation"][1][headerIndex])
                    if(Edit_Distance_Ratio(Attribute_Name,AtrName)<0.3):
                    #DMA_Result[table] = (DMA_Score, Attribute_Name, tableID)
                        DMA_Result[table] = (DMA_Score)

            #Normalize DMA Scores:
            Total_Score=0
            for table in DMA_Result:
                Total_Score+=DMA_Result[table]
            
            for table in DMA_Result:
                DMA_Result[table] = DMA_Result[table]/Total_Score
                
            
            return DMA_Result
    
    
def DMA_Augmentation_By_Attribute_Name(ABA, keyList, Aug_Attribute_Name, Threshold):
    keyListMap = {}
    keyList2 = []
    # Cleansing the Key List + Unique
    for x in range(len(keyList)):
        keyListMap[string_Cleanse(keyList[x])] = keyList[x]
        keyList2.append(string_Cleanse(keyList[x]))

    queryKeys = []
    for item in keyList2:
        if (item not in queryKeys):
            queryKeys.append(item)

    # Cleansing the Attribute Name
    Orig_Aug_Attribute_Name = Aug_Attribute_Name
    Aug_Attribute_Name = string_Cleanse(Aug_Attribute_Name)

    '''
    for table in DMA_Result:
         print(table+":"+str(DMA_Result[table]))
    '''

    DMA_Result=DMA_MATCH(keyList, Aug_Attribute_Name, Threshold)


    Key_Names=[]
    WT_Corpus={}

    with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
        WT_Corpus = pickle.load(f3)

    #Populating Augmentation by Attribute Name Dictionary: ABA: (queryKey)->(Value_Collection)
    for table in DMA_Result:
        if(Aug_Attribute_Name == string_Cleanse(DMA_Result[table][1])):
            contributed=0
            orig_tableKeys=copy.deepcopy(WT_Corpus[table]["relation"][0])
            table_Keys = WT_Corpus[table]["relation"][0]
            for x in range(0, len(table_Keys)):
                table_Keys[x] = string_Cleanse(table_Keys[x])
            for queryKey in queryKeys:
                if (queryKey in table_Keys):
                    contributed = 1
                    value = ""
                    for x in range(0, len(table_Keys)):
                        if (table_Keys[x] == queryKey):
                            value = WT_Corpus[table]["relation"][1][x]
                            break

                    if (keyListMap[queryKey] not in ABA):
                        ABA[keyListMap[queryKey]] = (value, DMA_Result[table][0], DMA_Result[table][2])
                    else:
                        ABA[keyListMap[queryKey]] += (value, DMA_Result[table][0], DMA_Result[table][2])
            if(contributed==1):
                for key in orig_tableKeys:
                    if(key not in Key_Names):
                        Key_Names.append(key)

            orig_tableKeys = []


    return ABA

            #print("Key Attribute Names used in the tables which contributed to augmentation, Name(T.K) : "+str(Key_Names))

def DMA_Synonyms(queryKeys,Aug_Attribute_Name, Threshold):
    synonyms=[]
    #tfidf_vectorizer = TfidfVectorizer(use_idf=False, smooth_idf=False)
    tfidf_vectorizer = TfidfVectorizer(use_idf=False, smooth_idf=False)
    with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
        WT_Corpus = pickle.load(f3)

        with open(get_Constant("T2Syn_DMA_Pickle"), 'rb') as f3:
            Synonym_Indx = pickle.load(f3)
            DMA_Result={}
            DMA_Result = DMA_MATCH(queryKeys,Aug_Attribute_Name, Threshold)
            for tableKey in DMA_Result:
                if(string_Cleanse(Aug_Attribute_Name) == string_Cleanse(DMA_Result[tableKey][1])):
                    if(tableKey in Synonym_Indx ):
                        syns=Synonym_Indx[tableKey]
                        for syn_item in syns:
                            if(syn_item[1] >Threshold):
                                documents=[]
                                documents.append( WT_Corpus[tableKey]["textBeforeTable"]+WT_Corpus[tableKey]["textAfterTable"])
                                documents.append(WT_Corpus[syn_item[2]]["textBeforeTable"] + WT_Corpus[syn_item[2]]["textAfterTable"])
                                tfidf = tfidf_vectorizer.fit_transform(documents)
                                # print (tfidf_matrix.shape)
                                pairwise_sim_sparse = tfidf * tfidf.T
                                pairwise_similarity = pairwise_sim_sparse.toarray()
                                #if(pairwise_similarity[0][1]<0.95):
                                    #if(WT_Corpus[tableKey]["url"]!= WT_Corpus[syn_item[2]]["url"]):
                                synonyms.append((tableKey,syn_item[0], syn_item[2], syn_item[1]))

        #print (synonyms)

        UniqueAugAttributeNameCol = []
        for item in synonyms:
            if (item not in UniqueAugAttributeNameCol):
                UniqueAugAttributeNameCol.append(item)

        return synonyms

def DMA_Syn_ABA(queryKeys,Aug_Attribute_Name, Threshold):

    AugAttributeNameCol = []
    AugAttributeNameCol.append(Aug_Attribute_Name)
    # If using Synonym Attributes of the Attribute name is Enabled:
    # Populate the  AugAttributeNameCol with its synonyms
    #  Setting theSynonym Attributes
    synonyms = DMA_Synonyms(queryKeys,Aug_Attribute_Name, Threshold)
    for syn in synonyms:
        AugAttributeNameCol.append(syn[1])

    UniqueAugAttributeNameCol = []
    for item in AugAttributeNameCol:
        if (item not in UniqueAugAttributeNameCol):
            UniqueAugAttributeNameCol.append(item)

    ABA={}
    for attrib in UniqueAugAttributeNameCol:
        ABA = DMA_Augmentation_By_Attribute_Name(ABA, queryKeys,attrib,Threshold)
    print(
        "\n################## Augmentation For Attribute: '" + Aug_Attribute_Name + "' and its synonyms (<Augemnted value> , <DMA_Score> , <Table_ID>)#####################\n")

    for queryKey in ABA:
        print(queryKey + ":" + repr(ABA[queryKey]))

def DMA_Schema_Auto_Completion():


    #queryKeyName="state"
    Augemntation_Attributs = DMA_Attribute_Discovery(US_States,0.01,True)

    for attrib in Augemntation_Attributs:
        DMA_Syn_ABA(US_States,attrib,0.01)
'''
This is the implementaion of InfoGather, which is one
of the baselines used in our experimental evaluations.
'''


def test():

    attrib="Political Party"
    ABA={}
    DMA_Augmentation_By_Attribute_Name(ABA,US_States, attrib, 0.3)
    print(
        "\n################## Augmentation For Attribute: '" + attrib + "' (<Augemnted value> , <DMA_Score> , <Table_ID>)#####################\n")
    for queryKey in ABA:
        print(queryKey + ":" + repr(ABA[queryKey]))

def Prepare_Dataset(GraphEnabled=False):
    pool = Pool(processes=8)
    #Merge_Json_Files()
    #Deduplicate_High_Quality_Json_Tables()
    EAB_Gen()

    Create_WIK()
    Create_WIA()

    if(GraphEnabled):
        with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
            WT_Corpus = pickle.load(f3)
        Corpus_EAB_Table_Num = len(WT_Corpus)
        
        #Train the Classifier with 10% of Labelled data
        #Create_Labeling(int(Corpus_EAB_Table_Num/10),0.3)
        #Create_Labeling(23,0.1)
        
        #Creating The Features in a parallel manner:
        print("Feature Generation Started ...")
        
        
        pool.apply_async(Feature_Gen, ["Feature_URL_Pickle"])
        pool.apply_async(Feature_Gen, ["Feature_CONTEXT_Pickle"])
        pool.apply_async(Feature_Gen, ["Feature_TUPLE_Pickle"])
        pool.apply_async(Feature_Gen, ["Feature_TableContent_Pickle"])
        pool.apply_async(Feature_Gen, ["Feature_Table2Context_Pickle"])
        pool.apply_async(Feature_Gen, ["Feature_AttributeNames_Pickle"])
        pool.apply_async(Feature_Gen, ["Feature_ColValues_Pickle"])
        pool.apply_async(ColWidth_Feature_Gen,[])
        
        '''Feature_Gen("Feature_URL_Pickle")
        Feature_Gen("Feature_CONTEXT_Pickle")        
        Feature_Gen("Feature_TUPLE_Pickle")
        Feature_Gen("Feature_TableContent_Pickle")
        Feature_Gen("Feature_Table2Context_Pickle")
        Feature_Gen("Feature_AttributeNames_Pickle")        
        Feature_Gen("Feature_ColValues_Pickle")
        ColWidth_Feature_Gen()'''
        
        
        
        pool.close()
        pool.join()
        
        print("Feature Generation Finished ...")
        
        SMW_Graph_Creation()
        #with open(get_Constant('SMW_Graph_Pickle'), 'wb') as outfile:
            #pickle.dump({}, outfile)
        
        
        #In order to Omit those tuples from the same original table:
        Clean_SMW_Graph()
        
        
        Generate_Parrallel_PPR(0.3)
        Merge_PPR_To_FPPR()
        
        #Needed for Attribute Discovery
        Create_T2Syn_From_SMW(20,0.001)
        
        #In order to Clean those synonms from the same original table, and keep only one of them in the Syn List.
        Clean_T2Syn_From_SMW()
        
        
        Write_T2Syn_From_SMW_To_File()
        

def Write_Dict_W_Table_To_File(Dict,DictName):
    with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
        dict = pickle.load(f3)
        Textfile = open(
                "C:/Saeed_Local_Files/InfoGather/" + Dataset +"_"+DictName+".txt", "w", encoding="utf-8")
        sorted_x = sorted(Dict.items(), key=lambda item:item[1], reverse=True)
        line_Num=0
        for data in sorted_x:
            line_Num+=1
            Textfile.write(repr(line_Num)+"-Node: "+repr(data[0])+" : "+repr(data[1])+"\n")
            Textfile.write(JsonTable_To_String(dict[data[0]]))
            Textfile.write("##############################################\n")
            
            
        
        Textfile.close()


def Cluster_Discovered_Attributes(TSP_Score_Dict, Threshold):
    
    All_Attribs={}
    All_Attribs['All Attribs'] = Get_All_Attribute_Names(TSP_Score_Dict, Threshold)
    Write_Only_Dict_To_File(All_Attribs,"AD_All_Attribs")
    print("All_Attribs['All Attribs']: "+str(len(All_Attribs['All Attribs'])))
        
    SynSets = Get_Synonym_Sets(TSP_Score_Dict, Threshold) 
    Write_Sets_To_File(SynSets,"Syn_Sets")
    print("len(SynSets): "+str(len(SynSets)))
    
    SynSets = Agglomerative_Cluster(SynSets)
    Write_Sets_To_File(SynSets,"Agglomerative_Cluster")
    print("Agglomerative_Cluster: "+str(len(SynSets)))
    
    return SynSets


def Get_All_Attribute_Names(TSP_Score_Dict, Threshold):
    with open(get_Constant("T2Syn_From_SMW"), 'rb') as f3:
        T2Syn_dict = pickle.load(f3)
        Attrib_Set=[]
        for TSP_Table in TSP_Score_Dict:
            if(TSP_Score_Dict[TSP_Table]>Threshold):
                for Syn_Tuple in T2Syn_dict[TSP_Table]:
                   Attrib_Set.append((Syn_Tuple[1],Syn_Tuple[2],Syn_Tuple[0])) 
                
                
                
        Attrib_Set = FuzzyGroup(Attrib_Set)
        return Attrib_Set
              
    
    
    
def Get_Synonym_Sets(TSP_Score_Dict, Threshold):
     #with open(get_Constant("T2Syn_Direct"), 'rb') as f3:
     with open(get_Constant("T2Syn_From_SMW"), 'rb') as f3:
        T2Syn_dict = pickle.load(f3)
        SynSets=[]
        
        #RETRIEVING ALL THE SYNONYMS OF EACH OF THE TSP TABLES
        for TSP_Table in TSP_Score_Dict:
            if(TSP_Score_Dict[TSP_Table]>Threshold):
                Syns={}
                for Syn_Tuple in T2Syn_dict[TSP_Table]:
                    if(Syn_Tuple[1] not in Syns):
                        Syns[Syn_Tuple[1]] = (Syn_Tuple[2], Syn_Tuple[0])
                SynSets.append((Syns,TSP_Score_Dict[TSP_Table]))
            
        return SynSets

def Write_Sets_To_File(Set, Name):
    Textfile = open(
                "C:/Saeed_Local_Files/InfoGather/" + Dataset +"_"+Name+".txt", "w", encoding="utf-8")
    line_Num=0
    for key in Set:
     
        Textfile.write(repr(key[0])+" : "+repr(key[1])+"\n")

        Textfile.write("##############################################\n")
        
        
    
    Textfile.close()
    
    
def Write_Only_Dict_To_File(Dict, Name):
    Textfile = open(
                "C:/Saeed_Local_Files/InfoGather/" + Dataset +"_"+Name+".txt", "w", encoding="utf-8")
    line_Num=0
    for key in Dict:
     
        Textfile.write(repr(key)+" :\n"+repr(Dict[key])+"\n")

        Textfile.write("##############################################\n")
        
        
    
    Textfile.close()

#def FuzzyCluster(SynSets):
    #return SynSets


def Attribute_discovery_Holistic(keyList, SeedThreshold):
    Seed_Tables = AD_Seed_Tables(keyList, SeedThreshold )
    print("Seed_Tables")
    print(Seed_Tables)
    print("size of Seed_Tables is: "+repr(len(Seed_Tables)))
    Write_Dict_W_Table_To_File(Seed_Tables,"AD_seeds")
    print("###############")
    TSP_Threshold = SeedMinTSPScore(Seed_Tables, SeedThreshold)
    TSP_Score_Dict = Compute_TSP(Seed_Tables, TSP_Threshold)
    print("TSP_Score_Dict")
    print(TSP_Score_Dict)
    Write_Dict_W_Table_To_File(TSP_Score_Dict,"AD_TSP_Score_Dict")
    print("###############")
    print("size of TSP_Score_Dict is: "+repr(len(TSP_Score_Dict)))
    
    #Write_Dict_To_File(TSP_Score_Dict,"TSP_Score_Dict")
    
    return Cluster_Discovered_Attributes(TSP_Score_Dict, TSP_Threshold)
    
    
def ABA_Prediction(keyList, TSP_Score_Dict):
    keyListMap = {}
    keyList2 = []
    # Cleansing the Key List + Unique
    for x in range(len(keyList)):
        keyListMap[string_Cleanse(keyList[x])] = keyList[x]
        keyList2.append(string_Cleanse(keyList[x]))

    queryKeys = []
    for item in keyList2:
        if (item not in queryKeys):
            queryKeys.append(item)
    
    with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
        WT_Corpus = pickle.load(f3)
        
        
    ABA={}
        
    #Populating Augmentation by Attribute Name Dictionary: ABA: (queryKey)->(Value_Collection)
    for table in TSP_Score_Dict:
            contributed=0
            orig_tableKeys=copy.deepcopy(WT_Corpus[table]["relation"][0])
            table_Keys = WT_Corpus[table]["relation"][0]
            for x in range(0, len(table_Keys)):
                table_Keys[x] = string_Cleanse(table_Keys[x])
            for queryKey in queryKeys:
                #if (queryKey in table_Keys):
                    #contributed = 1
                    value = ""
                    for x in range(0, len(table_Keys)):
                        if (Edit_Distance_Ratio(table_Keys[x],queryKey)<0.3):
                            
                            value = WT_Corpus[table]["relation"][1][x]
                            if (keyListMap[queryKey] not in ABA):
                                ABA[keyListMap[queryKey]] = [(value, TSP_Score_Dict[table], table)]
                            else:
                                ABA[keyListMap[queryKey]].append((value, TSP_Score_Dict[table], table))
                            break
            
    #FuzzyGrouping Step: Based  on Edit Distance
    #FuzzyGrouped_ABA={}
    #for key in ABA:
        #FuzzyGrouped_ABA[key] = FuzzyGroup(ABA[key])
    
    
    
    #Sort the predictions for each query key based on TSP score: It is done in FuzzyGroup Function
    #for key in FuzzyGrouped_ABA:
        #FuzzyGrouped_ABA[key] = sorted(FuzzyGrouped_ABA[key], key=lambda item:item[1], reverse=True)
        
    return ABA
      

def FuzzyGroup(InputList):
    FuzzyGroupedList=[]
    while (len(InputList)>0):
        NewGroupCentoid = string_Cleanse(InputList[0][0])
        NewGroupNames = [InputList[0][0]]
        NewGroupScore = InputList[0][1]
        NewGroupTables = [InputList[0][2]]
        del InputList[0]
        i=0
        while (i<len(InputList)):
            if (Edit_Distance_Ratio(string_Cleanse(InputList[i][0]),NewGroupCentoid)<0.3):
                if (InputList[i][0] not in NewGroupNames ):
                    NewGroupNames.append(InputList[i][0])
                NewGroupScore += InputList[i][1]
                if (InputList[i][2] not in NewGroupTables ):
                    NewGroupTables.append(InputList[i][2])
                del InputList[i]
            else:
                i+=1
        
        FuzzyGroupedList.append((NewGroupNames, NewGroupScore, NewGroupTables))
        
    

    FuzzyGroupedList = sorted(FuzzyGroupedList, key=lambda item:item[1], reverse=True)
        
    return FuzzyGroupedList


def ABA_Holilstic(keyList, SeedThreshold, AtrName):
    print("ABA Holistic For: "+AtrName)
    Seed_Tables = ABA_Seed_Tables(keyList, AtrName, SeedThreshold)
    print("Seed_Tables")
    print(Seed_Tables)
    print("size of Seed_Tables is: "+repr(len(Seed_Tables)))
    #Write_Dict_W_Table_To_File(Seed_Tables,"ABA_seeds")
    print("###############")
    TSP_Score_Dict = Compute_TSP(Seed_Tables, 0.01)
    print("TSP_Score_Dict")
    print(TSP_Score_Dict)
    #Write_Dict_W_Table_To_File(TSP_Score_Dict,"ABA_TSP_Score_Dict")
    print("###############")
    print("size of TSP_Score_Dict is: "+repr(len(TSP_Score_Dict)))
    ABA = ABA_Prediction(keyList, TSP_Score_Dict)
    #Write_ABA_Prediction_To_Excel(ABA)
    #Write_Only_Dict_To_File(ABA,"ABA_Holistic")
    return ABA
    

def Write_ABA_Prediction_To_Excel(ABA_Dict):
    New_WB = openpyxl.Workbook()
    New_WS = New_WB.active
    New_WS.title = 'Experiments'
    
    GT_WB = openpyxl.load_workbook(get_Constant("Search_GroundTruth_In_Dataset"))
    GT_WS = GT_WB.get_sheet_by_name('Search_GroundTruth_In_Dataset')
    
    #Copy InputKeys and the Ground Truth in Dataset to the Worksheet
    for i in range(1, GT_WS.max_row+1):
        #Input Keys:L
        New_WS.cell(row=i, column=1).value = GT_WS.cell(row=i, column=1).value 
        
        #Complete Ground Truth
        New_WS.cell(row=i, column=2).value = GT_WS.cell(row=i, column=2).value
        
        #Ground Truth in Dataset
        New_WS.cell(row=i, column=3).value = GT_WS.cell(row=i, column=3).value 
    
    
    
    New_WS.cell(row=1, column=4).value = 'Experiment_Result'
    New_WS.cell(row=1, column=4).font = Font(bold=True)
    
    
    for key in ABA_Dict:
        for i in range(2, GT_WS.max_row):
            if (New_WS.cell(row=i, column=1).value == key):
                New_WS.cell(row=i, column=4).value = repr(ABA_Dict[key])
                
    
    New_WB.save(get_Constant("Experiments")) 

 
    
def Search_GroundTruth_In_Dataset():
    wb = openpyxl.load_workbook(get_Constant("GroundTruth"))
    sheet = wb.get_sheet_by_name('GroundTruth')
    
    sheet.title = 'Search_GroundTruth_In_Dataset'
    
    sheet.cell(row=1, column=3).value = 'Data in Dataset'
    sheet.cell(row=1, column=3).font = Font(bold=True)
    
    
    
    
    #GroundTruth_dict={}
    #for i in range(2, sheet.max_row):
        #GroundTruth_dict[sheet.cell(row=i, column=1).value] = sheet.cell(row=i, column=2).value
        
    Search_GT_Dict={}
    with open(get_Constant("EAB_Merged_Pickle_URL"), 'rb') as f3:
        WT_Corpus = pickle.load(f3)
        cnt=0
        for tablekey in WT_Corpus:
            cnt+=1
            print(cnt)
            table = WT_Corpus[tablekey]["relation"]
            for i in range(len(table[0])):
                for j in range(2, sheet.max_row+1):
                    key = string_Cleanse(sheet.cell(row=j, column=1).value)
                    value = string_Cleanse(sheet.cell(row=j, column=2).value)
                    if(((Edit_Distance_Ratio(string_Cleanse(table[0][i]), key)<0.3) and (Edit_Distance_Ratio(string_Cleanse(table[1][i]), value)<0.3)) or ((Edit_Distance_Ratio(string_Cleanse(table[1][i]), key)<0.3) and (Edit_Distance_Ratio(string_Cleanse(table[0][i]), value)<0.3)) ):
                        
                        #Populating the Search_GT_Dict
                        if (tablekey not in Search_GT_Dict):
                            Search_GT_Dict[tablekey] = [(table[0][i],table[1][i])]
                        else:
                            Search_GT_Dict[tablekey].append((table[0][i],table[1][i]))
                            
                            
                        if(sheet.cell(row=j, column=3).value is None):
                            sheet.cell(row=j, column=3).value = "("+repr(table[0][i])+","+repr(table[1][i])+","+repr(tablekey)+") - "
                        else:
                            sheet.cell(row=j, column=3).value += "("+repr(table[0][i])+","+repr(table[1][i])+","+repr(tablekey)+") - "
                        
    wb.save(get_Constant("Search_GroundTruth_In_Dataset"))
    


    #Writing the Search_Gt_Dict to File
    Textfile_All_Tables = open(
                    "C:/Saeed_Local_Files/InfoGather/" + Dataset + "Tables_Containing_GroundTruth.txt", "w", encoding="utf-8")
    line_Num=0 
    for data in Search_GT_Dict:
        line_Num+=1
        Textfile_All_Tables.write(str(line_Num)+"-Lable: "+str(data)+"\n")
        Textfile_All_Tables.write(str(Search_GT_Dict[data]))
        Textfile_All_Tables.write("\nTable Content:\n")
        Textfile_All_Tables.write(JsonTable_To_String(WT_Corpus[data]))
        Textfile_All_Tables.write("##############################################\n")

        
        
    
    Textfile_All_Tables.close()

def hi():
    os.makedirs(get_Constant("FPPR")+"_test")




def Merge_ABA_Results(Aggregated_Col, PartialResult):
    #The results of each attribute should be merged with the Aggregated Column 
    
    for key in PartialResult:
        if(key not in Aggregated_Col):
            Aggregated_Col[key]=PartialResult[key]
        else:
            for elem in PartialResult[key]:
                Aggregated_Col[key].append(elem)    
            
        
def ValueFusion(Aggregated_Col) :
    #After all partial results are retrieved, we fuse the result list for each key to exctract only one result for each key
    
    FuzzyGrouped_Col={}
    
    for key in Aggregated_Col:
        value_list = []
        for elem in Aggregated_Col[key]:
            value_list.append(elem[0])
        FuzzyGrouped_Col[key] = Octopus.Octopus_FuzzyGroup(value_list, Octopus_Value_Dist_Threshold)
        
    # Depricated: Sort the predictions for each query key based on TSP score: It is done in FuzzyGroup Function
    #for key in FuzzyGrouped_Col:
        #FuzzyGrouped_Col[key] = sorted(FuzzyGrouped_Col[key], key=lambda item:item[1], reverse=True)[0][0][0] #First Attribute of the Attribute list as Centorid

    #output the canonical value of each cluster
    for key in FuzzyGrouped_Col:
        fuzzy_grouped_val = copy.copy(FuzzyGrouped_Col[key])
        clusterVal_list=[]
        for cluster in fuzzy_grouped_val:
            clusterVal_list.append(cluster[0][0])
        FuzzyGrouped_Col[key] = clusterVal_list
    
    return FuzzyGrouped_Col


def Extend_Table(queryKeys):
    ###
    # This function extends the query keys by utilizing Attribute Discovery (AD) and Augmentation By Attribute name (ABA).
    # Note that for debugging purposes, a set of values are assigned for each cell in the result table, that the first value 
    # (which is the most frequent value correponding to the query value) is the final value for that cell. 
    ###
    #First Attribute Discovery
    #Then, for each SynSet, running ABA and Aggregating the results.
    
    #Retrieving All the Synonym sets of Attribute Discovery Operation
    AD_Threshold=0.3
    SynSets = Attribute_discovery_Holistic(queryKeys, AD_Threshold)
    
    #The Final Extended Table
    Extended_Table=[]
    Extended_Table.append(queryKeys)
    
    queryKeysCopy=queryKeys.copy()
    
    ABA_Threshold = 0.3
    
    cntr=0
    #Column Extension for each Synonym set:
    for SynSet in SynSets:
        cntr+=1
        #if(cntr>3):
            #break
        Aggregated_Col={}
        print(SynSet)
        for atr in SynSet[0]:
            print(atr)
            #Partial Results Contain the list of result values with their attched scores for each attribute in Synset
            PartialResult = ABA_Holilstic(queryKeysCopy, ABA_Threshold, atr)
            
            #The results of each attribute should be merged with the Aggregated Column 
            Merge_ABA_Results(Aggregated_Col, PartialResult)
        
        #After all partial results are retrieved, we fuse the result list for each key to exctract only one result for each key
        Fused_Col = ValueFusion(Aggregated_Col)
        
        #Add_Fused_Column: 
        New_Col=[]
        for key in queryKeys:
            if(key in Fused_Col):
                New_Col.append(Fused_Col[key])
            else:
                New_Col.append("")
                
        Extended_Table.append(New_Col)
        
    cntr=0
    for col in Extended_Table:
        cntr+=1
        col.insert(0,["atomic-value",str(cntr)])
        
    outFile= "C:/Saeed_Local_Files/InfoGather/" + Dataset +"_TableExtension_Result.csv"
    print("write table into "+str(outFile))
    Write_Table_To_CSV(Extended_Table, outFile)
    
    return Extended_Table
        

if  (__name__ == "__main__"):
    #DMA_Schema_Auto_Completion()
    #test()
    #print(DMA_Synonyms(US_States,"Political Party",0.3))
    #compare_Three_Webtables('967-2', '2576-1')
    #(DMA_Syn_ABA(US_States, "Political Party", 0.3))
    #DMA_Attribute_Discovery(US_States, 0.6, Synonym_Enabled=False)
    #Create_T2Syn_DMA(Threshold=0.1)
    #Create_T2Syn_DMA(Threshold=0.1 )
    #Fetch_T2Syn("Player")
    #InsertData()
    
    #Create_WIK()
    #Create_WIA()
    #Feature_Gen("Feature_ColWidth_Pickle")


    #Prepare_Dataset(True)
    #Generate_Parrallel_PPR(0.3)
    #Merge_PPR_To_FPPR()
        
    #Needed for Attribute Discovery
    
    #Create_T2Syn_From_SMW(10,0.01)
    #Clean_T2Syn_From_SMW()
    #Write_T2Syn_From_SMW_To_File()
        
        
    #Write_Dict_To_File(AD_Seed_Tables(US_States, 0.3 ),"seeds")
    
    #Attribute_discovery_Holistic(Populate_QueryKeys_From_ExcelFile("C:/Saeed_Local_Files/T2D_233/Country.xlsx"), 0.3)
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/All-Countries.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/countries-rand10.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/countries-rand10.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/countries-rand10.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/companies-rand10.xlsx"
    input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/films-rand10.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/games-rand10.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/languages-rand10.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/Languages.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/cities-rand10.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/currencies-rand10.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/test1.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/All-Countries.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/Currency-All.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/Languages.xlsx"
    #input_query_file = "C:/Saeed_Local_Files/TestDataSets/T2D_233/sample_queries/Cities.xlsx"
    
    #print(Populate_QueryTable_From_ExcelFile(input_query_file)[0])
    Extend_Table(Populate_QueryTable_From_ExcelFile(input_query_file)[0])
    
    #EAB_Gen()
    
    #Needed for Attribute Discovery
    #Create_T2Syn_From_SMW(20,0.01)
        
    #In order to Clean those synonms from the same original table:
    #Clean_T2Syn_From_SMW()
    #Write_T2Syn_From_SMW_To_File()
    
    #ABA_Holilstic(Populate_QueryKeys(), 0.001,"album")

    #a=[("salam", 0.2, "1-0"),("salam",0.3,"1-1"),("borobaba",0.3,"2-0"),("hichi",0.4,"3-0"),("salam",0.2,"1-3"),("borobaba", 0.3,"2-1")]
    #print(FuzzyGroup(a))
    
    #print(Edit_Distance_Ratio("The Dark Knight Rises", "The Dark Knight"))
    #print(Populate_QueryKeys())
    #JsonToPickle()
    #Search_GroundTruth_In_Dataset()
    #Create_Labeling(10,0.3)
    #SMW_Graph_Creation()
    #Create_T2Syn_DMA(Threshold=0.7)
    
    #Generate_Parrallel_PPR(0.1)
    #Merge_PPR_To_FPPR()
        
    #Generate_Parrallel_PPR(0.96)
    #erge_PPR_To_FPPR()
        
        #Needed for Attribute Discovery
    #Create_T2Syn_From_SMW(20,0.001)
    #Merge_Json_Files()
    #Deduplicate_High_Quality_Json_Tables()
    #Filter_HighQuality_Json_Tables()
    #EAB_Gen()
